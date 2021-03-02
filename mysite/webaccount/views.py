import requests, json
from django.shortcuts import render,redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.models import User
from django.contrib import messages
from django.urls import reverse
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.admin.views.decorators import staff_member_required
from django.core.mail import EmailMessage
from django.utils import timezone
from django.contrib.sites.shortcuts import get_current_site
from django.db.models import Q
from .filters import *
from .forms import *
from .models import *
from django.http import HttpResponse
from openpyxl import load_workbook , Workbook # TODO: pip install openpyxl
# ? For Charts in openpyxl
from openpyxl.chart import BarChart, Reference, Series
from django.http import JsonResponse
import encodings
from django.core.paginator import Paginator
from datetime import datetime, timedelta

# ! Things to be noted:
# request.user.is_superuser == "False" => Relational Manager
# request.user.is_superuser == "True" => Admin
# request.user.is_superuser == "None" => Consultant


# TODO: Read Excel booksheet to plot charts
def read_workbook(x, wb=None):
    P1 = []
    for row in x.iter_rows(min_row=6, max_col=2, max_row=len(tuple(x.rows))):
        P1.append({row[0].value : row[1].value})    
    return P1

    

@login_required
def profile_view(request):
    if request.user.is_superuser == "False":
        return render(request, "webaccount/profile_rm.html", {})
    else:
        return render(request, 'webaccount/profile.html', {})




# Relational Manager View
@login_required
def RM_Dashboard(request):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        template_name="webaccount/rm_dashboard.html"
        context = {

        }
        return render(request, template_name, context)


# Relational Manager Client List View
@login_required
def Client_List(request):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        numer_of_results = None
        template_name="webaccount/rm_client_list.html"
        client_list_obj = Client_Personal_Info.objects.filter(managerRelational = User.objects.get(username=request.user.username))
        Client_Personal_Info_filter = Client_Personal_InfoFilter(request.GET, queryset=client_list_obj)
        context={
            'client_list_section' : True,
            'filter': Client_Personal_Info_filter,
            'numer_of_results' : numer_of_results

        }
        return render(request, template_name, context)


# All Clients Invoice List
@login_required
def client_invoice_list(request):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        number_of_results = None
        template_name="webaccount/rm_client_invoice_list.html"
        client_invoice_list_obj = Client_Personal_Info.objects.filter(managerRelational = User.objects.get(username=request.user.username))
        Client_Invoice_Info_filter = Client_Personal_InfoFilter(request.GET, queryset=client_invoice_list_obj)
        context={
            'client_invoice_list_section' : True,
            'filter': Client_Invoice_Info_filter,
            'numer_of_results' : number_of_results
        }
        return render(request, template_name, context)


# Relational Manager Reports List View
@login_required
def Reports_List(request):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        numer_of_results = None
        template_name="webaccount/rm_reports_list.html"
        RMReport_list_obj = RMReport.objects.all()
        RMReport_filter = RMReportFilter(request.GET, queryset=RMReport_list_obj)
        # print(client_list_obj, Client_Personal_Info_filter)
        context={
            'client_list_section' : True,
            'filter': RMReport_filter,
            'numer_of_results' : numer_of_results

        }
        return render(request, template_name, context)


# Specific Client Report Chart  List
@login_required
def Client_Report_Charts(request, client_id, report_id):
    P1 = None
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        template_name="webaccount/rm_charts.html"
        try:
            
            client = Client_Personal_Info.objects.get(id = client_id)
            report = RMReport.objects.get(id = report_id, client=client)
            wb2 = load_workbook(report.uploadFile)
            P1= read_workbook(wb2.active, wb2)
        except Client_Personal_Info.DoesNotExist:
            messages.success(request, "Invalid Client ID")
            return redirect(reverse("Client_List_URL"))
        except RMReport.DoesNotExist:
            messages.success(request, "Invalid Report ID")
            return redirect(reverse("Client_List_URL"))    
        except:
            messages.success(request, "Invalid Request")
            return redirect(reverse("Client_List_URL"))
        context={
            "client" : client,
            "report":report,
            'obj' :wb2,
            "P1" : P1
        }
        return render(request, template_name, context)


# Save  Generated Charts for client reports
def save_report_chart(request, client_id, report_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        if request.method == "POST" and request.is_ajax:
            try:
                client = Client_Personal_Info.objects.get(id = client_id)
                report = RMReport.objects.get(id = report_id, client=client)
                # print("==============================================================")
                # print(request.POST)
                # # print(len(request.POST["labels[]"]))
                # # print(JsonResponse.serialize(request.POST))
                # # for i in request.POST:
                # #     print(request.POST[i])
                # # print()
                # print("==============================================================")
                if request.POST["chart"] == '1':
                    report.is_saved_chart_1 = True
                    report.chartLabel_1 = request.POST["labels"]
                    report.chartData_1 = request.POST['dataValues']
                    report.chart_1_type = request.POST['chart-type']
                elif request.POST['chart'] == "2":
                    report.is_saved_chart_2 = True
                    report.chartLabel_2 = request.POST["labels"]
                    report.chartData_2 = request.POST['dataValues']
                    report.chart_2_type = request.POST['chart-type']
                elif request.POST['chart'] == "3":
                    report.is_saved_chart_3 = True
                    report.chartLabel_3 = request.POST["labels"]
                    report.chartData_3 = request.POST['dataValues']
                    report.chart_3_type = request.POST['chart-type']
                elif request.POST['chart'] == "4":
                    report.is_saved_chart_4 = True
                    report.chartLabel_4 = request.POST["labels"]
                    report.chartData_4 = request.POST['dataValues']
                    report.chart_4_type = request.POST['chart-type']
                elif request.POST['chart'] == "5":
                    report.is_saved_chart_5 = True
                    report.chartLabel_5 = request.POST["labels"]
                    report.chartData_5 = request.POST['dataValues']
                    report.chart_5_type = request.POST['chart-type']
                elif request.POST['chart'] == "6":
                    report.is_saved_chart_6 = True
                    report.chartLabel_6 = request.POST["labels"]
                    report.chartData_6 = request.POST['dataValues']
                    report.chart_6_type = request.POST['chart-type']
                report.save()
                # print(len(report.chartData_6), len(report.chartLabel_6))
                return JsonResponse(json.loads( json.dumps( {"success": True})), status=200)
            except Exception as e:
                print(e)
                return JsonResponse( json.loads( json.dumps( {"success": False})),status=400)    
        else:
            return JsonResponse( json.loads( json.dumps( {"success": False})),status=400)

@login_required
def Add_New_RMReport(request, client_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            client = Client_Personal_Info.objects.get(id = client_id)

        except:
            messages.success(request, "Invalid Client ID")
            return redirect(reverse("Client_List_URL"))
        if request.method != "POST":
            form  = RMReportForm(initial = {'client' : client})
        else:
            form = RMReportForm(request.POST, request.FILES)
            if form.is_valid():
                obj = form.save(commit = False)
                obj.client = client
                obj.save()
                messages.success(request, "Report has been added successfully.")
                return redirect(reverse("Client_Reports_URL", args = [obj.client.id]))
        context={
            'form' : form,
            'client' : client
        }
        return render(request, "webaccount/rm_new_report.html", context)

@login_required
def update_RMReport(request, report_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            report = RMReport.objects.get(id = report_id)
            obj = report
            if request.method != "POST":
                form = RMReportForm(instance = report)
            else:
                form = RMReportForm(request.POST, request.FILES, instance = report)
                if form.is_valid():
                    obj = form.save(commit = False)
                    obj.client = report.client
                    obj.save()
                    messages.success(request, "Report has been updated successfully.")
                    return redirect(reverse("Client_Reports_URL", args = [obj.client.id]))
            context={
                'form':form,
                'report' : report,
                'client' : obj.client
            }
            return render(request, "webaccount/rm_update_report.html", context)
        except Exception as e:
            print(e)
            messages.success(request, "Invalid Report ID")
            return redirect(reverse("Client_List_URL"))


@login_required
def Client_Reports(request, client_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            template_name="webaccount/rm_specific_reports_list.html"
            client = Client_Personal_Info.objects.get(id = client_id)
            RMReport_list_obj = RMReport.objects.filter(client = client)
            numer_of_results = None
            RMReport_filter = RMReportFilter(request.GET, queryset=RMReport_list_obj)
            context={
                'client_list_section' : True,
                'filter': RMReport_filter,
                'numer_of_results' : numer_of_results,
                'client' : client
            }
            return render(request, template_name, context)
        except Exception as e:
            messages.success(request, "Invalid Client ID")
            return redirect(reverse("Client_List_URL"))

# Specific Client Invoice List
@login_required
def invoice_list(request, obj_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            template_name="webaccount/rm_specific_client_invoice_list.html"
            client = Client_Personal_Info.objects.get(id = obj_id)
            invoices = RMInvoice.objects.filter(client=client)
            invoices_filter = RMInvoice_Filter(request.GET, queryset=invoices)
            numer_of_results = None
            context={
                'client_invoice_list_section' : True,
                'numer_of_results' : numer_of_results,
                'filter' : invoices_filter,
                'client' : client
            }
            return render(request, template_name, context)
        except Exception as e:
            print(e)
            messages.success(request, "Invalid Client ID")
            return redirect(reverse("Client_Invoice_List_URL"))

# Add Specific Client Invoice Entry
@login_required
def add_specific_client_invoice_entry(request, client_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            client = Client_Personal_Info.objects.get(id=client_id)
            if request.method != "POST":
                form = RMInvoiceForm()
            else:
                form = RMInvoiceForm(request.POST, request.FILES)
                if form.is_valid():
                    obj = form.save(commit = False)
                    obj.client = client
                    obj.statusType = "New"
                    obj.save()
                    if request.POST['comment'].strip() != "" or len(request.PSOT['comment'].strip() != 0):
                        RMInvoiceComment.objects.create(
                            user = User.objects.get(username = request.user.username),
                            invoice = obj,
                            comment = request.POST.get("comment","")
                        )
                    messages.success(request, "Successfully added invoice entry")
                    return redirect(reverse("Specific_Client_Invoice_URL", args=[ client_id]))
        except Exception as e:
            print(e)
            messages.success(request, "Invalid Client ID")
            return redirect(reverse("Client_Invoice_List_URL"))
        template_name="webaccount/rm_add_specific_client_invoice.html"
        context = {
            'client' : client,
            'form' : form
        }
        return render(request, template_name, context)

# READ SPECIFIC CLIENT INVOICE ENTER By Relational Manager
@login_required
def read_specfic_client_invoice_entry(request, client_id, invoice_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            client = Client_Personal_Info.objects.get(id = client_id)
            invoice = RMInvoice.objects.get(client = client, id=invoice_id)
            comments = invoice.rminvoicecomment_set.all()
        except Exception as e:
            print(e)
            messages.success(request, "Invalid client ID or Invoice ID")
            return redirect(reverse("Client_Invoice_List_URL"))
        template_name = "webaccount/read_specific_client_invoice.html"
        context = {
            'client' : client,
            'obj' : invoice,
            'comments' : comments
        }
        return render(request, template_name, context)


# READ SPECIFIC CLIENT INVOICE ENTER  By CLIENT
@login_required
def read_specfic_client_invoice_entry_client_view(request, client_id, invoice_id):
    if request.user.is_superuser == "True"  :
        return redirect("profile_url")
    else:
        try:
            client = Client_Personal_Info.objects.get(id = client_id)
            invoice = RMInvoice.objects.get(client = client, id=invoice_id)
            comments = invoice.rminvoicecomment_set.all()
        except Exception as e:
            print(e)
            messages.success(request, "Invalid client ID or Invoice ID")
            return redirect(reverse("Client_Invoice_List_URL"))
        template_name = "webaccount/update_specific_client_invoice.html"
        context = {
            'client' : client,
            'obj' : invoice,
            'comments' : comments
        }
        return render(request, template_name, context)


@login_required
def update_file_again_by_client(request, client_id, invoice_id):
    if request.user.is_superuser == "True"  :
        return redirect("profile_url")
    else:
        if request.method != "POST":
            # return redirect("profile_url")
            return redirect(reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_CLIENT_VIEW_URL", args=[client_id, invoice_id]))
        else:
            # print(request.FILES)
            try:
                client = Client_Personal_Info.objects.get(id = client_id)
                invoice = RMInvoice.objects.get(client = client, id=invoice_id)
                invoice.uploadFile = request.FILES['uploadFile']
                invoice.statusType = "New"
                invoice.save()
                # EMAIL HANDLER
                domain =  get_current_site(request)
                url =  reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_URL", args=[client_id, invoice_id])
                build_link = str(request.scheme) + str("://") + str(domain) + str(url)
                mail_subject = "Invoice document has bee updated by : " + str(client.Email)
                message_title = "An email has been recieved from the Django Admin Group of Companies\n\n"
                message = message_title + "Please visit the following link to get updates about invoice\n\n\n"
                message += build_link
                to_email = str(client.managerRelational.email)
                email = EmailMessage(mail_subject, message, to=[to_email])
                email.send()
                messages.success(request, "Invioce Document/Image  has been updated again!")
                return redirect(reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_CLIENT_VIEW_URL", args=[client_id, invoice_id]))
            except Exception as e:
                print(e)
                messages.success(request, "Invalid client ID or Invoice ID")
                return redirect(reverse("Client_Invoice_List_URL"))
            # template_name = "webaccount/update_specific_client_invoice.html"
            # context = {
            #     'client' : client,
            #     'obj' : invoice
            # }
            # return render(request, template_name, context)


# UPDATE SPECIFIC CLIENT INVOICE ENTRY
@login_required
def update_specific_client_invoice_entry(request, client_id, invoice_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            if request.method == "POST":
                # print("****************************")
                # print(request.POST)
                # print("****************************")
                client = Client_Personal_Info.objects.get(id=client_id)
                invoice = RMInvoice.objects.get(client= client, id=invoice_id)
                invoice.statusType = request.POST.get("statusType", "New")
                invoice.lastUpdate = timezone.now()
                invoice.save()

                # EMAIL HANDLER
                domain =  get_current_site(request)
                url =  reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_CLIENT_VIEW_URL", args=[client_id, invoice_id])
                build_link = str(request.scheme) + str("://") + str(domain) + str(url)
                mail_subject = "Invoice has bee updated to : " + invoice.statusType
                message_title = "An email has been recieved from the Django Admin Group of Companies\n\n"
                message = message_title + "Please visit the following link to get updates about your invoice\n\n\n"
                message += build_link
                to_email = str(client.Email)
                email = EmailMessage(mail_subject, message, to=[to_email])
                email.send()
                messages.success(request, "Staus has been updated.")
                return redirect(reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_URL", args=[client_id,  invoice_id]))
            else:
                messages.success(request, "Invalid client ID or Invoice ID")
                return redirect(reverse("Client_Invoice_List_URL"))    
        except Exception as e:
            print(e)
            messages.success(request, "Invalid client ID or Invoice ID")
            return redirect(reverse("Client_Invoice_List_URL"))
        # template_name = "webaccount/update_specific_client_invoice.html"
        # context = {
        #     'client' : client,
        #     'obj' : invoice
        # }
        # return render(request, template_name, context)


# UPDATE SPECIFIC CLIENT INVOICE ENTRY COMMENT
@login_required
def update_specific_client_invoice_entry_comments(request, client_id, invoice_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            if request.method == "POST":
                # print("****************************")
                # print(request.POST)
                # print("****************************")
                client = Client_Personal_Info.objects.get(id=client_id)
                invoice = RMInvoice.objects.get(client= client, id=invoice_id)
                # invoice.comment = request.POST.get("comment", invoice.comment)
                RMInvoiceComment.objects.create(
                    user = User.objects.get(username = request.user.username),
                    invoice = invoice,
                    comment = request.POST.get("comment","")
                )
                invoice.lastUpdate = timezone.now()
                invoice.save()
                domain =  get_current_site(request)
                url =  reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_CLIENT_VIEW_URL", args=[client_id, invoice_id])
                build_link = str(request.scheme) + str("://") + str(domain) + str(url)
                mail_subject = "A comment has been added from Relational Manager to the Invoice #" + str(invoice.id)
                message_title = "An email has been recieved from the Django Admin Group of Companies\n\n"
                message = message_title + "Please visit the following link to get updates about your invoice\n\n\n"
                message += build_link
                to_email = str(client.Email)
                email = EmailMessage(mail_subject, message, to=[to_email])
                email.send()
                messages.success(request, "Comment has been send")
                return redirect(reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_URL", args=[client_id,  invoice_id]))
            else:
                messages.success(request, "Invalid client ID or Invoice ID")
                return redirect(reverse("Client_Invoice_List_URL"))    
        except Exception as e:
            print(e)
            messages.success(request, "Invalid client ID or Invoice ID")
            return redirect(reverse("Client_Invoice_List_URL"))


# ! COmment Added by Cient onto an invoice
@login_required
def update_specific_client_invoice_entry_client_comments(request, client_id, invoice_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            if request.method == "POST":
                # print("****************************")
                # print(request.POST)
                # print("****************************")
                client = Client_Personal_Info.objects.get(id=client_id)
                invoice = RMInvoice.objects.get(client= client, id=invoice_id)
                # invoice.comment = request.POST.get("comment", invoice.comment)
                RMInvoiceComment.objects.create(
                    user = User.objects.get(username = request.user.username),
                    invoice = invoice,
                    comment = request.POST.get("comment","")
                )
                invoice.lastUpdate = timezone.now()
                invoice.save()
                domain =  get_current_site(request)
                url =  reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_URL", args=[client_id, invoice_id])
                build_link = str(request.scheme) + str("://") + str(domain) + str(url)
                mail_subject = "A comment has been added from client : " + str(client.Email)  + "to the Invoice #" + str(invoice.id)
                message_title = "An email has been recieved from the Django Admin Group of Companies\n\n"
                message = message_title + "Please visit the following link to get updates about invoice\n\n\n"
                message += build_link
                to_email = str(client.managerRelational.email)
                email = EmailMessage(mail_subject, message, to=[to_email])
                email.send()
                messages.success(request, "Comment has been send")
                return redirect(reverse("READ_SPECFIC_CLIENT_INVOICE_ENTRY_CLIENT_VIEW_URL", args=[client_id,  invoice_id]))
            else:
                messages.success(request, "Invalid client ID or Invoice ID")
                return redirect(reverse("Client_Invoice_List_URL"))    
        except Exception as e:
            print(e)
            messages.success(request, "Invalid client ID or Invoice ID")
            return redirect(reverse("Client_Invoice_List_URL"))


@login_required
def Read_Client_Report(request, client_id, report_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            n = 0
            template_name="webaccount/rm_specific_reports_list_view.html"
            client = Client_Personal_Info.objects.get(id = client_id)
            report = RMReport.objects.get(client = client, id=report_id)
            wb2 = load_workbook(report.uploadFile).active
            try:
                n = len(wb2[1])
            except:
                n = 0
            context={
                'client' : client,
                'report' : report,
                'obj' :wb2,
                'n': n
            }
            return render(request, template_name,context)
        except Exception as e:
            print(e)
            messages.success(request, "Invalid Document Read Request")
            return redirect(reverse("Client_List_URL"))

@login_required
def Client_Notify(request, client_id, report_id):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        try:
            client = Client_Personal_Info.objects.get(id = client_id)
            report = RMReport.objects.get(client = client, id=report_id)
            report.status = "Confirmed"
            report.save()
            domain =  get_current_site(request)
            url =  reverse("Reports_List_Update_URL", args=[report.id])
            build_link = str(request.scheme) + str("://") + str(domain) + str(url)
            mail_subject = "Report Confirmation"
            message_title = "An email has been recieved from the Django Admin Group of Companies\n\n"

            message = message_title + "Please visit the following link to get updates about your report confirmation request\n\n\n"
            message += build_link
            to_email = str(client.Email)
            email = EmailMessage(mail_subject, message, to=[to_email])
            email.send()
            messages.success(request, "Client has been notify about report confirmation")
            return redirect(reverse("Client_Reports_URL", args=[report.client.id]))
        except Exception as e:
            print(e)
            messages.success(request, "Invalid Requested Page")
            return redirect(reverse("Client_List_URL"))

# Relational Manager Invoices List View
@login_required
def Invoices_List(request):
    if request.user.is_superuser == "True":
        return redirect("profile_url")
    else:
        template_name="webaccount/rm_invoices_list.html"
        context={

        }
        return render(request, template_name, context)


def index_view(request):
    # if request.user.is_authenticated and request.user.is_superuser:
    if request.user.is_authenticated:
        if request.user.is_superuser == "False" :
            return redirect("RM_Dashboard_URL")
        return redirect(reverse('profile_url'))
    else:
        if request.method != "POST":
            return render(request, 'webaccount/login.html', {})
        # elif request.method == "GET":
        else:
        # if authenticate
            try:
                user = authenticate(username=request.POST['username'], password = request.POST['password'])
            except:
                messages.error(request, "Incorrect username or password.")
            if user is not None:
                if user.is_active:
                    login(request, user)
                    # print("***************************************")
                    # print(request.user.is_superuser)
                    # print("***************************************")
                    if request.user.is_superuser == "False" :
                        return redirect("RM_Dashboard_URL")
                    else:
                        return redirect(reverse('profile_url'))
                else:
                    messages.error(request, "User account has been deactivate")
            else:
                if request.POST['username'] == '' and request.POST['password'] == '':
                    messages.error(request, "Both username and password are empty.")
                elif request.POST['username'] == '':
                    messages.error(request, "Username is empty.")
                elif request.POST['password'] == '':
                    messages.error(request, "Password is empty.")
                else:
                    messages.error(request, "Incorrect username or password.")
        return render(request, 'webaccount/login.html', {})

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, "User has been logout successfully.")
    return redirect(reverse("index_url"))


@login_required()
def change_password(request):
    if request.method != 'POST':
        form = PasswordChangeForm(user=request.user)
    else:
        form = PasswordChangeForm(data=request.POST, user=request.user)
        if form.is_valid():
            form.save()
            update_session_auth_hash(request, form.user)
            messages.success(request, 'Password has been updated.')
            if request.user.is_superuser:
                logout(request)
                return redirect(reverse('admin:login'))
            else:
                logout(request)
                return redirect(reverse('index_url'))
    if request.user.is_superuser == "False" :
        return render(request, 'webaccount/change_password_rm.html', {'form': form, 'section': "editProfile"})
    return render(request, 'webaccount/change_password.html', {'form': form, 'section': "editProfile"})

@login_required()
def edit_profile(request):
    if request.method!='POST':
        form = EditProfileForm(instance = request.user)
    else:
        form = EditProfileForm(request.POST, instance = request.user)
        if form.is_valid():
            form.save()
            # return redirect('profile', username = request.user.username)
            # return HttpResponseRedirect(reverse('profile', args=[request.user.username]))
            if request.user.is_superuser == "False":
                messages.success(request, "Profie has beed updated.")
                return redirect(reverse("RM_Dashboard_URL"))
            messages.success(request, "Profie has beed updated.")
            return redirect(reverse('index_url'))
    if request.user.is_superuser == "False":
        return render(request, 'webaccount/edit_profile_rm.html',{'form' : form})
    return render(request, 'webaccount/edit_profile.html',{'form' : form})

@login_required()
@staff_member_required
def send_quote_view(request, num):
    try:
        client = Client_Personal_Info.objects.get(id = num)
        d = client.clientrequireddocuments_set.all()
        e = Required_Documents.objects.all()
    except Client_Personal_Info.DoesNotExist:
        return redirect(reverse('admin:index'))
    context={
        'client': client,
        'title_site' : "Send Quote",
        'title' : "Client '" + str(client.Name) + "' Details",
        'documents': e,
        'd' : d
    }
    return render(request, "webaccount/send_quote_clinet.html",context)


@login_required()
@staff_member_required
def send_quote_mail_view(request, client_id):
    A = list(dict(request.POST).keys())
    if "csrfmiddlewaretoken" in  A:
        A.pop(A.index("csrfmiddlewaretoken"))
    if "documents" in A:
        A.pop(A.index("documents"))
    try:
        client = Client_Personal_Info.objects.get(id =client_id)
        if client.status == "Pending" :
            mail_subject = "Resend Quote Email"
            message_title = "An email has been recieved from the Django Admin Group of Companies"
            message_subject = "The documents need to be re-submitted"
            if len(A) != 0:
                message = "{title}\n{subject}\n{list}\n".format(title = message_title, subject = message_subject, list=A)
            else:
                message = "A resend email has been recieved from the Django Admin Group of Companies. You account has been activated."
            messages.success(request, 'A resend email has been send to the client to inform him about his status regarding account')
        else:
            mail_subject = "Quote Email"
            message_title = "An email has been recieved from the Django Admin Group of Companies"
            message_subject = "The documents need to be submitted"
            if len(A) != 0:
                message = "{title}\n{subject}\n{list}\n".format(title = message_title, subject = message_subject, list=A)
            else:
                message = "An email has been recieved from the Django Admin Group of Companies.You account has been activated."
            messages.success(request, 'An email has been send to the client to inform him about his status regarding account')
        client.status = "Pending"
        client.save()
        to_email = str(client.Email)
    except:
        messages.error(request, 'Client Does Not Exist.')
        return redirect(reverse("admin:index"))
    email = EmailMessage(mail_subject, message, to=[to_email])
    email.send()
    return redirect(reverse("send_quote_specific", args=[client_id]))


@login_required
@staff_member_required
def change_document_submitted_status_aaproved(request, client_id, document_id):
    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        document = ClientRequiredDocuments.objects.get(client = client, id =document_id)
        document.status = "Approved"
        document.save()
        messages.success("Status has been set to Approved.")
        return redirect(reverse("send_quote_specific", args=[client_id]))
    except:
        messages.error(request, 'Invalid Document.')
        return redirect(reverse("send_quote_specific", args=[client_id]))

@login_required
@staff_member_required
def change_document_submitted_status_rejected(request, client_id, document_id):
    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        document = ClientRequiredDocuments.objects.get(client = client, id =document_id)
        document.status = "Rejected"
        document.delete()
        messages.success("Status has been set to Rejected.")
        return redirect(reverse("send_quote_specific", args=[client_id]))
    except:
        messages.error(request, 'Invalid Document.')
        return redirect(reverse("send_quote_specific", args=[client_id]))

@login_required
@staff_member_required
def sendConsultantRequestQuote(request, client_id, consultant_request_id):
    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        consultant_request = ConsulatationRequest.objects.get(client= client, id=consultant_request_id)
        # return redirect(reverse("admin:webaccount_consulatationrequest_changelist"))
        template_name="webaccount/send_consultant_quote_clinet.html"
        context={
            'client': client,
            'title_site' : "Send Consultant Quote",
            'title' : "Client '" + str(client.Name) + "' Consultant",
            'consultant_request': consultant_request,
            'form' : ConsultantRequestUpdateForm(instance = consultant_request),
            'form_2' : FeedbackForm()
        }
        return render(request,
                    template_name,
                    context
                )
    except Exception as e:
        messages.success(request,"Invlid Request")
        return redirect(reverse("admin:webaccount_consulatationrequest_changelist"))

# Save and Send
@login_required
@staff_member_required
def save_send_quote_consultant_mail_view(request, client_id, consultant_request_id):
    try:
        client = Client_Personal_Info.objects.get(id =client_id)
        consultant_request = ConsulatationRequest.objects.get(client=client, id=consultant_request_id)
        form = ConsultantRequestUpdateForm(request.POST, instance=consultant_request)
        try:
            if form.is_valid():
                data_1, data_2  = form.cleaned_data['price'], form.cleaned_data['clientPaidAllAmount']
                obj = form.save(commit=False)
                obj.update_timestamp =  timezone.now()
                obj.save()
            else:
                messages.success(request, (form.errors))
                return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))
        except Exception as e:
            messages.success(request, str(e))
            return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))
        domain = current_site = get_current_site(request)
        url =  reverse("sendConsultantRequestQuote_URL", args=[client.id, consultant_request.id])
        build_link = str(request.scheme) + str("://") + str(domain) + str(url)
        mail_subject = "Consultation Quote Email [{status}]".format(status = obj.status)
        message_title = "An email has been recieved from the Django Admin Group of Companies\n\n"
        message_subject = "Consultant Quote"
        message = message_title + "Please visit the following link to get updates about your consultation request\n\n\n"
        message += build_link
        to_email = str(client.Email)
        email = EmailMessage(mail_subject, message, to=[to_email])
        if obj.status != "New":
            messages.success(request, 'A resend email has been send to the client to inform him about his status regarding consultant request.')
        else:
            messages.success(request, 'An email has been send to the client to inform him about his status regarding consultant request and client consultation quote status has been changed from "New" to "Pending."')
        if obj.status == "New":
            obj.status = "Pending"
            obj.save()
        email.send()
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))
    except Exception as e:
        print(e)
        messages.error(request, 'Client Does Not Exist.')
        return redirect(reverse("admin:index"))




# Rejected
@login_required
@staff_member_required
def reject_quote_consultant_mail_view(request, client_id, consultant_request_id):
    if request.method == "POST":
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id, consultant_request_id]))
    else:
        try:
            client = Client_Personal_Info.objects.get(id =client_id)
            consultant_request = ConsulatationRequest.objects.get(client=client, id=consultant_request_id)
            consultant_request.status = "Rejected"
            consultant_request.save()
            domain = current_site = get_current_site(request)
            url =  reverse("sendConsultantRequestQuote_URL", args=[client.id, consultant_request.id])
            build_link = str(request.scheme) + str("://") + str(domain) + str(url)
            # -------------
            mail_subject = "Consultation Quote Email [Rejected]"
            message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
            message =  message_title + "Please visit the following link to get updates about your consultation request\n\n\n"
            message += build_link
            messages.success(request, 'An email has been send to the client to inform him about his status regarding consultant request.')
            to_email = str(client.Email)
        except:
            messages.error(request, 'Client Does Not Exist.')
            return redirect(reverse("admin:index"))
        email = EmailMessage(mail_subject, message, to=[to_email])
        email.send()
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))

@staff_member_required
@login_required
def confirm_quote_consultant_mail_view(request, client_id ,consultant_request_id):
    if request.method == "POST":
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id, consultant_request_id]))
    else:
        try:
            client = Client_Personal_Info.objects.get(id =client_id)
            consultant_request = ConsulatationRequest.objects.get(client=client, id=consultant_request_id)
            if consultant_request.clientPaidAllAmount:
                consultant_request.status = "Confirmed"
                consultant_request.save()
                domain = current_site = get_current_site(request)
                url =  reverse("sendConsultantRequestQuote_URL", args=[client.id, consultant_request.id])
                build_link = str(request.scheme) + str("://") + str(domain) + str(url)
                # -------------
                mail_subject = "Consultation Quote Email [Confirmed]"
                message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
                message_subject = "Consultant Quote"
                message =  message_title + "Please visit the following link to get updates about your consultation request\n\n\n"
                message += build_link
                messages.success(request, 'An email has been send to the client to inform him about his status regarding consultant request.')
                to_email = str(client.Email)
            else:
                messages.success(request, "You can't confirm the consultation request because the client didn't pay yet.")
                return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))
        except:
            messages.error(request, 'Client Does Not Exist.')
            return redirect(reverse("admin:index"))
        email = EmailMessage(mail_subject, message, to=[to_email])
        email.send()
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))


# CLose
@login_required
@staff_member_required
def  close_quote_consultant_mail_view(request, client_id , consultant_request_id):
    if request.method == "POST":
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id, consultant_request_id]))
    else:
        try:
            client = Client_Personal_Info.objects.get(id =client_id)
            consultant_request = ConsulatationRequest.objects.get(client=client, id=consultant_request_id)
            consultant_request.status = "Close"
            consultant_request.save()
            domain = current_site = get_current_site(request)
            url =  reverse("sendConsultantRequestQuote_URL", args=[client.id, consultant_request.id])
            build_link = str(request.scheme) + str("://") + str(domain) + str(url)
            # -------------
            mail_subject = "Consultation Quote Email [Close]"
            message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
            message_subject = "Consultant Quote"
            message =  message_title + "Please visit the following link to get updates about your consultation request\n\n\n"
            message += build_link
            messages.success(request, 'An email has been send to the client to inform him about his status regarding consultant request.')
            to_email = str(client.Email)
        except:
            messages.error(request, 'Client Does Not Exist.')
            return redirect(reverse("admin:index"))
        email = EmailMessage(mail_subject, message, to=[to_email])
        email.send()
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))



# Complete
@login_required
@staff_member_required
def  complete_quote_consultant_mail_view(request, client_id , consultant_request_id):
    if request.method == "POST":
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id, consultant_request_id]))
    else:
        try:
            client = Client_Personal_Info.objects.get(id =client_id)
            consultant_request = ConsulatationRequest.objects.get(client=client, id=consultant_request_id)
            consultant_request.status = "Completed"
            consultant_request.save()
            domain = current_site = get_current_site(request)
            url =  reverse("sendConsultantRequestQuote_URL", args=[client.id, consultant_request.id])
            build_link = str(request.scheme) + str("://") + str(domain) + str(url)
            # -------------
            mail_subject = "Consultation Quote Email [Close]"
            message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
            message_subject = "Consultant Quote"
            message =  message_title + "Please visit the following link to get updates about your consultation request\n\n\n"
            message += build_link
            messages.success(request, 'An email has been send to the client to inform him about his status regarding consultant request.')
            to_email = str(client.Email)
        except:
            messages.error(request, 'Client Does Not Exist.')
            return redirect(reverse("admin:index"))
        email = EmailMessage(mail_subject, message, to=[to_email])
        email.send()
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))

@login_required
@staff_member_required
def sendFile(request,client_id ,consultant_request_id):
    if request.method == "GET":
        return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))
    else:
        try:
            client = Client_Personal_Info.objects.get(id =client_id)
            consultant_request = ConsulatationRequest.objects.get(client=client, id=consultant_request_id)
            form = FeedbackForm(request.POST, request.FILES, instance=consultant_request)
            if form.is_valid():
                form.save()
                messages.success(request, "File has been send successffully")
                return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))
            else:
                messages.success(request, str(form.errors))
                return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))
        except Exception as e:
            print(e)
            messages.success(request, "There is an error while uploading documents")
            return redirect(reverse("sendConsultantRequestQuote_URL", args=[client_id,consultant_request_id ]))


@login_required
@staff_member_required
def declineVIew(request, client_id, consultant_request_id):
    A = [
        'Price is High',
        "I don't need the consultation",
        'The time is not suitable for me',
        'Other',
    ]
    if request.method != "POST":
        messages.success(request, 'Invalid Request.')
        return redirect(reverse(
            "sendConsultantRequestQuote_URL", args=[
            client_id,
            consultant_request_id
        ]))
    else:
        try:
            client = Client_Personal_Info.objects.get(id =client_id)
            consultant_request = ConsulatationRequest.objects.get(client=client, id=consultant_request_id)
            consultant_request.status = "Declined"
            if int(request.POST['customRadio']) != 4:
                consultant_request.decline_explanation = A[int(request.POST['customRadio'])]
            else:
                consultant_request.decline_explanation = request.POST['customRadio1']
            consultant_request.save()
            print(request.POST)
            messages.success(request, 'Consultation has been declined')
            return redirect(reverse(
                "sendConsultantRequestQuote_URL", args=[
                client_id,
                consultant_request_id
            ]))
        except Exception as e:
            messages.success(request, 'Invalid Request')
            return redirect(reverse(
                "sendConsultantRequestQuote_URL", args=[
            client_id,
            consultant_request_id
            ]))

@login_required
@staff_member_required
def ratingsView(request, client_id, consultant_request_id):
    if request.method != "POST":
        messages.success(request, 'Invalid Request.')
        return redirect(reverse(
            "sendConsultantRequestQuote_URL", args=[
            client_id,
            consultant_request_id
        ]))
    else:
        try:
            client = Client_Personal_Info.objects.get(id =client_id)
            consultant_request = ConsulatationRequest.objects.get(client=client, id=consultant_request_id)
            consultant_request.rating = int(request.POST['whatever1'])
            consultant_request.save()
            messages.success(request, 'Ratings has been submitted.')
            return redirect(reverse(
                "sendConsultantRequestQuote_URL", args=[
                client_id,
                consultant_request_id
            ]))
        except Exception as e:
            messages.success(request, 'Invalid Request')
            return redirect(reverse(
                "sendConsultantRequestQuote_URL", args=[
            client_id,
            consultant_request_id
            ]))

#  --------------------------------------------------------------
#  Pickup Order requests

@login_required
@staff_member_required
def viewPickUpRequest(request, client_id, pickup_order_id):

    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        order_Request = PickUpRequestOrders.objects.get(id=pickup_order_id)
        template_name = "webaccount/pickup_order_requests_list.html"
        context = {
            'client' : client,
            'order_Request' : order_Request,
            'title_site' : "Pick Up Order Details",
            'title' : "Client '" + str(client.Name) + "'s Pickup Order Details"
        }
        return render(request, template_name, context)
    except Exception as e:
        print(e)
        messages.success(request,"Invlid Request")
        return redirect(reverse("admin:webaccount_pickuprequestorders_changelist"))


@login_required
@staff_member_required
def viewPickUpRequestAccept(request, client_id, pickup_order_id):
    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        order_Request = PickUpRequestOrders.objects.get(id=pickup_order_id)
        if order_Request.shippingMethod is None:
            messages.success(request,"Shipping method field is required")
            return redirect(reverse("viewPickUpRequest_URL", args= [
                client_id ,
                pickup_order_id
            ]))
        else:
            order_Request.status="ACCEPT"
            order_Request.updated_timestamp = timezone.now()
            order_Request.save()
            domain = current_site = get_current_site(request)
            url =  reverse("viewPickUpRequest_URL", args=[client.id, order_Request.id])
            build_link = str(request.scheme) + str("://") + str(domain) + str(url)
            # -------------
            mail_subject = "Pickup Order Request Email [Accepted]"
            message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
            message_subject = "Pickup Order Request"
            message =  message_title + "Please visit the following link to get updates about your Pickup Order Request\n\n\n"
            message += build_link
            messages.success(request, 'An email has been send to the client to inform him about his status regarding Pickup Order Request.')
            to_email = str(client.Email)
            email = EmailMessage(mail_subject, message, to=[to_email])
            email.send()
            return redirect(reverse("viewPickUpRequest_URL", args= [
                client_id ,
                pickup_order_id
            ]))
    except Exception as e:
        print(e)
        messages.success(request,"Invlid Request")
        return redirect(reverse("admin:webaccount_pickuprequestorders_changelist"))



@login_required
@staff_member_required
def viewPickUpRequestReject(request, client_id, pickup_order_id):
    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        order_Request = PickUpRequestOrders.objects.get(id=pickup_order_id)
        order_Request.status="REJECT"
        order_Request.updated_timestamp = timezone.now()
        order_Request.save()
        domain = current_site = get_current_site(request)
        url =  reverse("viewPickUpRequest_URL", args=[client.id, order_Request.id])
        build_link = str(request.scheme) + str("://") + str(domain) + str(url)
        # -------------
        mail_subject = "Pickup Order Request Email [Rejected]"
        message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
        message_subject = "Pickup Order Request"
        message =  message_title + "Please visit the following link to get updates about your Pickup Order Request\n\n\n"
        message += build_link
        messages.success(request, 'An email has been send to the client to inform him about his status regarding Pickup Order Request.')
        to_email = str(client.Email)
        email = EmailMessage(mail_subject, message, to=[to_email])
        email.send()
        return redirect(reverse("viewPickUpRequest_URL", args= [
            client_id ,
            pickup_order_id
        ]))
    except Exception as e:
        print(e)
        messages.success(request,"Invlid Request")
        return redirect(reverse("admin:webaccount_pickuprequestorders_changelist"))



@login_required
@staff_member_required
def viewPickUpRequestOnDelivery(request, client_id, pickup_order_id):
    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        order_Request = PickUpRequestOrders.objects.get(id=pickup_order_id)
        order_Request.status="ON DELIVERY"
        order_Request.updated_timestamp = timezone.now()
        order_Request.save()
        domain = current_site = get_current_site(request)
        url =  reverse("viewPickUpRequest_URL", args=[client.id, order_Request.id])
        build_link = str(request.scheme) + str("://") + str(domain) + str(url)
        # -------------
        mail_subject = "Pickup Order Request Email [On Delivery]"
        message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
        message_subject = "Pickup Order Request"
        message =  message_title + "Please visit the following link to get updates about your Pickup Order Request\n\n\n"
        message += build_link
        messages.success(request, 'An email has been send to the client to inform him about his status regarding Pickup Order Request.')
        to_email = str(client.Email)
        email = EmailMessage(mail_subject, message, to=[to_email])
        email.send()
        return redirect(reverse("viewPickUpRequest_URL", args= [
            client_id ,
            pickup_order_id
        ]))
    except Exception as e:
        print(e)
        messages.success(request,"Invlid Request")
        return redirect(reverse("admin:webaccount_pickuprequestorders_changelist"))



@login_required
@staff_member_required
def viewPickUpRequestRecieved(request, client_id, pickup_order_id):
    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        order_Request = PickUpRequestOrders.objects.get(id=pickup_order_id)
        order_Request.status="RECEIVED"
        order_Request.updated_timestamp = timezone.now()
        order_Request.save()
        domain = current_site = get_current_site(request)
        url =  reverse("viewPickUpRequest_URL", args=[client.id, order_Request.id])
        build_link = str(request.scheme) + str("://") + str(domain) + str(url)
        # -------------
        mail_subject = "Pickup Order Request Email [Received]"
        message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
        message_subject = "Pickup Order Request"
        message =  message_title + "Please visit the following link to get updates about your Pickup Order Request\n\n\n"
        message += build_link
        messages.success(request, 'An email has been send to the client to inform him about his status regarding Pickup Order Request.')
        to_email = str(client.Email)
        email = EmailMessage(mail_subject, message, to=[to_email])
        email.send()
        return redirect(reverse("viewPickUpRequest_URL", args= [
            client_id ,
            pickup_order_id
        ]))
    except Exception as e:
        messages.success(request,"Invlid Request")
        return redirect(reverse("admin:webaccount_pickuprequestorders_changelist"))


@login_required
@staff_member_required
def viewPickUpRequestOnFailed(request, client_id, pickup_order_id):
    try:
        client = Client_Personal_Info.objects.get(id = client_id)
        order_Request = PickUpRequestOrders.objects.get(id=pickup_order_id)
        order_Request.status="FAILED"
        order_Request.updated_timestamp = timezone.now()
        order_Request.save()
        domain = current_site = get_current_site(request)
        url =  reverse("viewPickUpRequest_URL", args=[client.id, order_Request.id])
        build_link = str(request.scheme) + str("://") + str(domain) + str(url)
        # -------------
        mail_subject = "Pickup Order Request Email [Failed to Pickup]"
        message_title = "An email has been recieved from the Django Admin Group of Companies\n\n\n"
        message =  message_title + "Please visit the following link to get updates about your Pickup Order Request\n\n\n"
        message += build_link
        messages.success(request, 'An email has been send to the client to inform him about his status regarding Pickup Order Request.')
        to_email = str(client.Email)
        email = EmailMessage(mail_subject, message, to=[to_email])
        email.send()
        return redirect(reverse("viewPickUpRequest_URL", args= [
            client_id ,
            pickup_order_id
        ]))
    except Exception as e:
        print(e)
        messages.success(request,"Invlid Request")
        return redirect(reverse("admin:webaccount_pickuprequestorders_changelist"))

@login_required
@staff_member_required
def SeeDetails(request):
    template_name='see_details.html'
    context = {
        "objects" : ParentModel.objects.all()
    }
    return render(request, template_name, context)


@login_required
def Client_API(request):
    template_name="webaccount/client_api.html"
    doc_list = []
    context = {
        'doc_list' : doc_list
    }
    return render(request, template_name, context)


@login_required
def client_specific_api(request, invoice_id):
    template_name="webaccount/client_api_invoice_details.html"
    doc_list = []
    context={
        'doc_list' : doc_list
    }
    return render(request, template_name, context)

# ? Admin : New Package 
@login_required
def NEW_PACKAGE(request):
    if request.user.is_superuser != "True":
        messages.success(request, "Only Administration have access to it. Thanks!")
        return redirect(reverse("profile_url"))
    else:
        template_name=  "webaccount/package/add.html"
        if request.method != "POST":
            form = PackageForm()
        else:
            form = PackageForm(request.POST)
            if form.is_valid():
                form.save()
                messages.success(request,"Package has been addedd successfully.")
                return redirect(reverse("PACKAGE_LIST_URL"))
        context = {
            "form" : form
        }
        return render(request, template_name, context)


# ? Admin:  Detail Package
@login_required
def DETAIL_PACKAGE(request, package_id):
    if request.user.is_superuser != "True":
        messages.success(request, "Only Administration have access to it. Thanks!")
        return redirect(reverse("profile_url"))
    else:
        try:
            obj = PackageModel.objects.get(id = package_id)
            template_name = "webaccount/package/detail.html"
            context = {
                "obj" : obj
            }
            return render(request, template_name, context)
        except  Exception as  e:
            # print(e)
            messages.success(request,"Invalid Package-ID or request")
            return redirect(reverse("PACKAGE_LIST_URL"))


# ? Admin : Edit Package 
@login_required
def EDIT_PACKAGE(request, package_id):
    if request.user.is_superuser != "True":
        messages.success(request, "Only Administration have access to it. Thanks!")
        return redirect(reverse("profile_url"))
    else:
        try:
            obj = PackageModel.objects.get(id = package_id)
            template_name=  "webaccount/package/edit.html"
            if request.method != "POST":
                form = PackageForm(instance=obj)
            else:
                form = PackageForm(instance=obj, data = request.POST)
                if form.is_valid():
                    form.save()
                    messages.success(request,"Package has been updated successfully.")
                    return redirect(reverse("PACKAGE_LIST_URL"))
            context = {
                "form" : form,
                "obj" : obj
            }
            return render(request, template_name, context)
        except Exception  as e:
            messages.success(request, "Invalid Package ID")
            return redirect(reverse("PACKAGE_LIST_URL"))

# ? Admin: Delete Package 
@login_required
def DELETE_PACKAGE(request, package_id):
    if request.user.is_superuser != "True":
        messages.success(request, "Only Administration have access to it. Thanks!")
        return redirect(reverse("profile_url"))
    else:
        try:
            obj = PackageModel.objects.get(id = package_id)
            obj.delete()
            messages.success(request, "Package has been deleted successfullly!")
            return redirect(reverse("PACKAGE_LIST_URL"))
        except Exception as e:
            messages.success(request, "There might be some problem during delete process. We shortly notifie you after configure it. Thanks")
            return redirect(reverse("PACKAGE_LIST_URL"))


# ? ADMIN: OBJECT LIST
@login_required
def PACKAGE_LIST(request):
    search  = None
    if request.user.is_superuser != "True":
        messages.success(request, "Only Administration have access to it. Thanks!")
        return redirect(reverse("profile_url"))
    else:
        template_name= "webaccount/package/objects_list.html"
        objs = PackageModel.objects.all()
        try:
            # ! Pagiation
            paginator  = Paginator(objs, 10) # Show 1 objects per page.
            page_number = request.GET.get('page', "1")
            page_obj = paginator.get_page(page_number)
            # ! Filter
            if request.GET.get("page"):
                package_filter = PackageModelFilter(request.GET, queryset=page_obj.object_list)
                search = None
            elif request.GET.get("id")  :
                package_filter = PackageModelFilter(request.GET, queryset=objs)
                search = True
            else:
                package_filter = PackageModelFilter(request.GET, queryset=page_obj.object_list)
                search = None
        except  Exception as  e:
            print("*****************************")
            print(e)
            print("*****************************")
        context  = {
            "objs" : objs,
            "filter" : package_filter,
            'page_obj': page_obj,
            "search"  :search
        }
        return render(request, template_name, context)



# ? Admin Dashboard Page
def ADMIN_DASHBOARD(request):
    # return HttpResponse("")
    # !!    Save Data related to sector and its corresponding amount
    sector_dict ={}
    vat_amount = 0
    book_amount = 0

    # Paid Amount By VAT:
    for i in Client_Personal_Info.objects.all().filter(Services="['VAT']"):
        vat_amount+= float(i.package_price)

    for i in Client_Personal_Info.objects.all().filter(Services="['BookKeeping']"):
        book_amount+= float(i.package_price)


    filter_parameter = None
    template_name=  "admin/dashboard.html"
    sector_objects = Sector.objects.all()
    for i in sector_objects:
        sector_dict[str(i.Name)] = float(i.get_paid_amount())
    d = timezone.now()-timedelta(days=14)
    users = User.objects.all().filter(last_login__lt = d)
    # print(users)
    # print(timezone.now()-timedelta(days=14))
    
    if request.GET.get("SERVICES", None):
        if request.GET.get("SERVICES", None) == "All" or request.GET.get("SERVICES", None) == "Both":
            Clients_Objects = Client_Personal_Info.objects.all()
            filter_parameter = request.GET.get("SERVICES", None)
        else:
            # print("['{}']".format(request.GET.get("SERVICES", None)))
            try:
                Clients_Objects = Client_Personal_Info.objects.all().filter(Services__icontains = (request.GET.get("SERVICES", None)))
                print(Clients_Objects)
                filter_parameter = request.GET.get("SERVICES", None)
            except :
                Clients_Objects = Client_Personal_Info.objects.all()
                filter_parameter = None
    else:
        Clients_Objects = Client_Personal_Info.objects.all()
        filter_parameter = None
    
    # ! Retrieve All  Clients  Objects
    context={
        "Clients_Objects" : Clients_Objects ,
        "sector_objects" : sector_objects,
        "New_Consultant_Manager_Counts" : ConsulatationRequest.new_consultant_requests.all().count(),
        "Confirmed_Consultant_Manager_Counts" : ConsulatationRequest.confirmed_consultant_requests.all().count(),
        "Pending_Consultant_Manager_Counts" : ConsulatationRequest.pending_consultant_requests.all().count(),
        "Completed_Consultant_Manager_Counts" : ConsulatationRequest.completed_consultant_requests.all().count(),
        "Close_Consultant_Manager_Counts" : ConsulatationRequest.close_consultant_requests.all().count(),
        "Rejected_Consultant_Manager_Counts" : ConsulatationRequest.rejected_consultant_requests.all().count(),
        "Declined_Consultant_Manager_Counts" : ConsulatationRequest.declined_consultant_requests.all().count(),
        "accept_pick_up_request_order_counts": PickUpRequestOrders.accept_pick_up_request_order.all().count(),
        "reject_pick_up_request_order_counts": PickUpRequestOrders.reject_pick_up_request_order.all().count(),
        "on_delivery_pick_up_request_order_counts": PickUpRequestOrders.on_delivery_pick_up_request_order.all().count(),
        "receive_pick_up_request_order_counts": PickUpRequestOrders.receive_pick_up_request_order.all().count(),
        "failed_pick_up_request_order_counts": PickUpRequestOrders.failed_pick_up_request_order.all().count(),
        "pending_pick_up_request_order_counts": PickUpRequestOrders.pending_pick_up_request_order.all().count(),
        'new_client_invoice_counts' : clientInvoice.new_client_invoice.all().count(),
        'approved_client_invoice_counts' : clientInvoice.approved_client_invoice.all().count(),
        'rejected_client_invoice_counts' : clientInvoice.rejected_client_invoice.all().count(),
        'processed_client_invoice_counts' : clientInvoice.processed_client_invoice.all().count(),
        'postponed_client_invoice_counts' : clientInvoice.postponed_client_invoice.all().count(),
        'New_Client_Personal_Info_Objects_counts'       :   Client_Personal_Info.New_Client_Personal_Info_Objects.all().count(),
        'Active_Client_Personal_Info_Objects_counts'        :   Client_Personal_Info.Active_Client_Personal_Info_Objects.all().count(),
        'Pending_Client_Personal_Info_Objects_counts'       :   Client_Personal_Info.Pending_Client_Personal_Info_Objects.all().count(),
        'Completed_Client_Personal_Info_Objects_counts'     :   Client_Personal_Info.Completed_Client_Personal_Info_Objects.all().count(),
        'Disabled_Client_Personal_Info_Objects_counts'      :   Client_Personal_Info.Disabled_Client_Personal_Info_Objects.all().count(),
        "users_count" : users.count(),
        "sector_dict" : sector_dict,
        "vat_amount": vat_amount,
        "book_amount": book_amount,
        "filter_parameter" : filter_parameter
    }
    return render(
        request,
        template_name,
        context
    )